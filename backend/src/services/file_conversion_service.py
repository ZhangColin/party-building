# -*- coding: utf-8 -*-
"""文件转换服务"""
import os
import subprocess
import tempfile
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class ConversionError(Exception):
    """文件转换错误"""
    pass


class FileConversionService:
    """文件转换服务"""

    EXTENSION_MAP = {
        "word": [".docx", ".doc"],
        "excel": [".xlsx", ".xls"],
        "pdf": [".pdf"],
        "wps": [".wps"],
        "markdown": [".md"],
        "text": [".txt"],
        "image": [".png", ".jpg", ".jpeg", ".gif", ".webp"],
    }

    def _get_file_type(self, filename: str) -> str | None:
        """根据文件名获取文件类型"""
        ext = Path(filename).suffix.lower()
        for file_type, extensions in self.EXTENSION_MAP.items():
            if ext in extensions:
                return file_type
        return None

    async def convert_to_markdown(self, file_path: str, file_type: str) -> str:
        """将文件转换为 Markdown 格式"""
        if file_type == "markdown":
            return Path(file_path).read_text(encoding="utf-8")
        elif file_type == "text":
            content = Path(file_path).read_text(encoding="utf-8")
            return f"```\n{content}\n```"
        elif file_type == "word":
            return await self._convert_word(file_path)
        elif file_type == "excel":
            return await self._convert_excel(file_path)
        elif file_type == "pdf":
            return await self._convert_pdf(file_path)
        elif file_type == "wps":
            return await self._convert_wps(file_path)
        elif file_type == "image":
            raise ConversionError("图片文件不支持转换")
        else:
            raise ConversionError(f"不支持的文件类型: {file_type}")

    async def _convert_word(self, file_path: str) -> str:
        """转换 Word 文档（支持 .docx 和 .doc）"""
        with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as tmp:
            tmp_path = tmp.name

        # 检查文件扩展名
        ext = Path(file_path).suffix.lower()

        if ext == ".doc":
            # 旧版 .doc 文件，需要先用 LibreOffice 转换为 docx
            try:
                with tempfile.TemporaryDirectory() as tmp_dir:
                    # 先转换为 docx
                    subprocess.run(
                        ["libreoffice", "--headless", "--convert-to", "docx",
                         "--outdir", tmp_dir, file_path],
                        capture_output=True,
                        timeout=30
                    )
                    # 找到转换后的 docx 文件
                    docx_files = list(Path(tmp_dir).glob("*.docx"))
                    if docx_files:
                        # 再用 pandoc 转换 docx 到 markdown
                        result = subprocess.run(
                            ["pandoc", "-f", "docx", "-t", "markdown",
                             "-o", tmp_path, str(docx_files[0])],
                            capture_output=True,
                            text=True,
                            timeout=30
                        )
                    else:
                        raise ConversionError("DOC 转换失败：未找到转换后的文件")
            except FileNotFoundError:
                raise ConversionError("DOC 转换失败：未找到 LibreOffice 工具，请安装 libreoffice")
        else:
            # .docx 文件直接用 pandoc 转换
            result = subprocess.run(
                ["pandoc", "-f", "docx", "-t", "markdown", "-o", tmp_path, file_path],
                capture_output=True,
                text=True,
                timeout=30
            )

        if result.returncode != 0:
            raise ConversionError(f"Word 文档转换失败: {result.stderr}")

        content = Path(tmp_path).read_text(encoding="utf-8")
        os.unlink(tmp_path)
        return content

    async def _convert_excel(self, file_path: str) -> str:
        """转换 Excel 文件"""
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True)
        lines = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"\n## {sheet_name}\n")
            for row in ws.iter_rows(values_only=True):
                row_text = "| " + " | ".join(str(cell) if cell is not None else "" for cell in row) + " |"
                lines.append(row_text)

        wb.close()
        return "\n".join(lines)

    async def _convert_pdf(self, file_path: str) -> str:
        """转换 PDF 文档"""
        with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as tmp:
            tmp_path = tmp.name

        # 使用 pandoc 转换 PDF
        result = subprocess.run(
            ["pandoc", "-f", "pdf", "-t", "markdown", "-o", tmp_path, file_path],
            capture_output=True,
            text=True,
            timeout=60
        )

        if result.returncode != 0:
            # 尝试使用 pdftotext 作为备选方案
            try:
                result = subprocess.run(
                    ["pdftotext", "-layout", file_path, tmp_path],
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                if result.returncode != 0:
                    raise ConversionError(f"PDF 转换失败: {result.stderr}")
            except FileNotFoundError:
                raise ConversionError("PDF 转换失败：未找到 pandoc 或 pdftotext 工具")

        content = Path(tmp_path).read_text(encoding="utf-8")
        os.unlink(tmp_path)
        return content

    async def _convert_wps(self, file_path: str) -> str:
        """转换 WPS 文档"""
        with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as tmp:
            tmp_path = tmp.name

        # WPS 文档可以用 LibreOffice 或 pandoc 转换
        # 首先尝试用 LibreOffice 转换为 docx，再用 pandoc 转换
        result = subprocess.run(
            ["pandoc", "-f", "docx", "-t", "markdown", "-o", tmp_path, file_path],
            capture_output=True,
            text=True,
            timeout=30
        )

        if result.returncode != 0:
            # 尝试使用 LibreOffice 转换
            try:
                with tempfile.TemporaryDirectory() as tmp_dir:
                    # 先转换为 docx
                    subprocess.run(
                        ["libreoffice", "--headless", "--convert-to", "docx",
                         "--outdir", tmp_dir, file_path],
                        capture_output=True,
                        timeout=30
                    )
                    # 找到转换后的 docx 文件
                    docx_files = list(Path(tmp_dir).glob("*.docx"))
                    if docx_files:
                        # 再用 pandoc 转换 docx 到 markdown
                        result = subprocess.run(
                            ["pandoc", "-f", "docx", "-t", "markdown",
                             "-o", tmp_path, str(docx_files[0])],
                            capture_output=True,
                            text=True,
                            timeout=30
                        )
                        if result.returncode != 0:
                            raise ConversionError(f"WPS 转换失败: {result.stderr}")
                    else:
                        raise ConversionError("WPS 转换失败：未找到转换后的文件")
            except FileNotFoundError:
                raise ConversionError("WPS 转换失败：未找到 pandoc 或 LibreOffice 工具")

        content = Path(tmp_path).read_text(encoding="utf-8")
        os.unlink(tmp_path)
        return content
